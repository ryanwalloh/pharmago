import time
import json
import csv
import io
import os
from datetime import datetime, timedelta
from django.db import models, connection, transaction
from django.apps import apps
from django.core.exceptions import ValidationError, ObjectDoesNotExist
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.db.models import Q, Count, Avg, Sum, Max, Min
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FileUploadParser
from rest_framework.throttling import UserRateThrottle, AnonRateThrottle
from openpyxl import Workbook, load_workbook

from api.core.permissions import IsOwnerOrReadOnly
from .models import SystemHealth, ApiUsage, GlobalSearchLog, BulkOperationLog
from .serializers import (
    SystemHealthSerializer, SystemHealthCreateSerializer, SystemHealthUpdateSerializer,
    SystemHealthListSerializer, SystemHealthDetailSerializer,
    ApiUsageSerializer, ApiUsageCreateSerializer, ApiUsageListSerializer, ApiUsageStatsSerializer
)


# ============================================================================
# SYSTEM HEALTH VIEWSET
# ============================================================================

class SystemHealthViewSet(viewsets.ModelViewSet):
    """ViewSet for system health monitoring and management"""
    
    queryset = SystemHealth.objects.all()
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['component', 'status']
    search_fields = ['component', 'message']
    ordering_fields = ['last_check', 'response_time', 'status']
    ordering = ['-last_check']
    
    def get_serializer_class(self):
        if self.action == 'create':
            return SystemHealthCreateSerializer
        elif self.action in ['update', 'partial_update']:
            return SystemHealthUpdateSerializer
        elif self.action == 'list':
            return SystemHealthListSerializer
        elif self.action == 'retrieve':
            return SystemHealthDetailSerializer
        return SystemHealthSerializer
    
    @action(detail=False, methods=['get'])
    def overall_status(self, request):
        """Get overall system health status"""
        try:
            # Check database connectivity
            with connection.cursor() as cursor:
                cursor.execute("SELECT 1")
                db_status = 'healthy'
                db_response_time = 0
        except Exception as e:
            db_status = 'critical'
            db_response_time = None
            
        # Check cache connectivity (if Redis is configured)
        try:
            from django.core.cache import cache
            cache.set('health_check', 'ok', 10)
            cache_result = cache.get('health_check')
            cache_status = 'healthy' if cache_result == 'ok' else 'warning'
        except Exception:
            cache_status = 'critical'
        
        # Check file storage
        try:
            media_root = os.path.join(os.getcwd(), 'media')
            os.makedirs(media_root, exist_ok=True)
            test_file = os.path.join(media_root, 'health_check.txt')
            with open(test_file, 'w') as f:
                f.write('health check')
            os.remove(test_file)
            storage_status = 'healthy'
        except Exception:
            storage_status = 'critical'
        
        # Determine overall status
        statuses = [db_status, cache_status, storage_status]
        if 'critical' in statuses:
            overall_status = 'critical'
        elif 'warning' in statuses:
            overall_status = 'warning'
        else:
            overall_status = 'healthy'
        
        # Update or create health records
        components = [
            ('database', db_status, 'Database connectivity check'),
            ('cache', cache_status, 'Cache connectivity check'),
            ('storage', storage_status, 'File storage check'),
            ('overall', overall_status, 'Overall system health')
        ]
        
        for component, status_val, message in components:
            SystemHealth.objects.create(
                component=component,
                status=status_val,
                message=message,
                response_time=db_response_time if component == 'database' else None
            )
        
        return Response({
            'overall_status': overall_status,
            'components': {
                'database': db_status,
                'cache': cache_status,
                'storage': storage_status
            },
            'timestamp': timezone.now().isoformat()
        })
    
    @action(detail=False, methods=['get'])
    def component_status(self, request):
        """Get status of specific system components"""
        component = request.query_params.get('component', 'overall')
        try:
            latest = SystemHealth.objects.filter(component=component).latest('last_check')
            return Response({
                'component': latest.component,
                'status': latest.status,
                'message': latest.message,
                'last_check': latest.last_check,
                'response_time': latest.response_time
            })
        except SystemHealth.DoesNotExist:
            return Response({'error': f'No status found for component: {component}'}, 
                          status=status.HTTP_404_NOT_FOUND)
    
    @action(detail=False, methods=['post'])
    def check_health(self, request):
        """Manually trigger health check for all components"""
        # This will trigger the overall_status action
        return self.overall_status(request)


# ============================================================================
# API USAGE VIEWSET
# ============================================================================

class ApiUsageViewSet(viewsets.ModelViewSet):
    """ViewSet for API usage tracking and analytics"""
    
    queryset = ApiUsage.objects.all()
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['user', 'endpoint', 'method', 'status_code']
    search_fields = ['endpoint', 'user_agent', 'ip_address']
    ordering_fields = ['timestamp', 'response_time', 'status_code']
    ordering = ['-timestamp']
    
    def get_serializer_class(self):
        if self.action == 'create':
            return ApiUsageCreateSerializer
        elif self.action == 'list':
            return ApiUsageListSerializer
        return ApiUsageSerializer
    
    @action(detail=False, methods=['get'])
    def stats(self, request):
        """Get API usage statistics"""
        # Get date range from query params
        days = int(request.query_params.get('days', 7))
        start_date = timezone.now() - timedelta(days=days)
        
        # Filter usage data
        usage_data = ApiUsage.objects.filter(timestamp__gte=start_date)
        
        # Calculate statistics
        total_requests = usage_data.count()
        average_response_time = usage_data.aggregate(avg_time=Avg('response_time'))['avg_time'] or 0
        success_rate = (usage_data.filter(status_code__lt=400).count() / total_requests * 100) if total_requests > 0 else 0
        
        # Top endpoints
        top_endpoints = usage_data.values('endpoint').annotate(
            count=Count('id')
        ).order_by('-count')[:10]
        
        # Top users
        top_users = usage_data.filter(user__isnull=False).values(
            'user__email'
        ).annotate(
            count=Count('id')
        ).order_by('-count')[:10]
        
        # Hourly distribution
        hourly_distribution = {}
        for hour in range(24):
            count = usage_data.filter(timestamp__hour=hour).count()
            hourly_distribution[f"{hour:02d}:00"] = count
        
        stats_data = {
            'total_requests': total_requests,
            'average_response_time': round(average_response_time, 2),
            'success_rate': round(success_rate, 2),
            'top_endpoints': list(top_endpoints),
            'top_users': list(top_users),
            'hourly_distribution': hourly_distribution,
            'period_days': days
        }
        
        serializer = ApiUsageStatsSerializer(stats_data)
        return Response(serializer.data)
    
    @action(detail=False, methods=['get'])
    def my_usage(self, request):
        """Get current user's API usage"""
        user_usage = ApiUsage.objects.filter(user=request.user).order_by('-timestamp')
        page = self.paginate_queryset(user_usage)
        if page is not None:
            serializer = ApiUsageListSerializer(page, many=True)
            return self.get_paginated_response(serializer.data)
        
        serializer = ApiUsageListSerializer(user_usage, many=True)
        return Response(serializer.data)


# ============================================================================
# GLOBAL SEARCH VIEWSET
# ============================================================================

class GlobalSearchViewSet(viewsets.ViewSet):
    """ViewSet for global search across all models"""
    
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['query']
    ordering_fields = ['timestamp']
    ordering = ['-timestamp']
    
    @action(detail=False, methods=['get'])
    def search(self, request):
        """Perform global search across all searchable models"""
        query = request.query_params.get('q', '')
        model_type = request.query_params.get('model', 'all')
        limit = int(request.query_params.get('limit', 50))
        
        if not query:
            return Response({'error': 'Search query is required'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        results = {}
        searchable_models = {
            'user': 'api.User',
            'customer': 'api.Customer', 
            'pharmacy': 'api.Pharmacy',
            'medicine': 'api.MedicineCatalog',
            'order': 'api.Order',
            'address': 'api.Address'
        }
        
        # Log the search
        GlobalSearchLog.objects.create(
            user=request.user,
            query=query,
            model_type=model_type,
            results_count=0
        )
        
        if model_type == 'all':
            models_to_search = searchable_models.values()
        elif model_type in searchable_models:
            models_to_search = [searchable_models[model_type]]
        else:
            return Response({'error': f'Invalid model type: {model_type}'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        for model_path in models_to_search:
            try:
                app_label, model_name = model_path.split('.')
                model = apps.get_model(app_label, model_name)
                
                # Get search fields for this model
                search_fields = getattr(model, 'SEARCH_FIELDS', [])
                if not search_fields:
                    # Default search fields
                    if hasattr(model, 'name'):
                        search_fields = ['name']
                    elif hasattr(model, 'title'):
                        search_fields = ['title']
                    elif hasattr(model, 'email'):
                        search_fields = ['email']
                    else:
                        continue
                
                # Build search query
                q_objects = Q()
                for field in search_fields:
                    q_objects |= Q(**{f"{field}__icontains": query})
                
                # Execute search
                model_results = model.objects.filter(q_objects)[:limit]
                
                if model_results.exists():
                    # Serialize results (basic serialization)
                    serialized_results = []
                    for obj in model_results:
                        serialized_obj = {
                            'id': obj.id,
                            'model': model_name.lower(),
                            'display_name': str(obj),
                            'url': f"/api/v1/{model_name.lower()}/{obj.id}/"
                        }
                        
                        # Add specific fields based on model
                        if hasattr(obj, 'name'):
                            serialized_obj['name'] = obj.name
                        if hasattr(obj, 'email'):
                            serialized_obj['email'] = obj.email
                        if hasattr(obj, 'phone'):
                            serialized_obj['phone'] = obj.phone
                        
                        serialized_results.append(serialized_obj)
                    
                    results[model_name.lower()] = {
                        'count': len(serialized_results),
                        'results': serialized_results
                    }
                    
            except Exception as e:
                # Log error but continue with other models
                continue
        
        # Update search log with results count
        total_results = sum(result['count'] for result in results.values())
        if GlobalSearchLog.objects.filter(user=request.user, query=query).exists():
            latest_log = GlobalSearchLog.objects.filter(user=request.user, query=query).latest('timestamp')
            latest_log.results_count = total_results
            latest_log.save()
        
        return Response({
            'query': query,
            'total_results': total_results,
            'results': results,
            'timestamp': timezone.now().isoformat()
        })
    
    @action(detail=False, methods=['get'])
    def suggestions(self, request):
        """Get search suggestions based on recent queries"""
        query = request.query_params.get('q', '')
        if not query:
            return Response([])
        
        # Get recent similar queries
        suggestions = GlobalSearchLog.objects.filter(
            query__icontains=query
        ).values('query').distinct().order_by('-timestamp')[:10]
        
        return Response([s['query'] for s in suggestions])


# ============================================================================
# BULK OPERATIONS VIEWSET
# ============================================================================

class BulkOperationsViewSet(viewsets.ViewSet):
    """ViewSet for bulk CRUD operations"""
    
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FileUploadParser]
    
    @action(detail=False, methods=['post'])
    def bulk_create(self, request):
        """Bulk create records"""
        model_type = request.data.get('model_type')
        data = request.data.get('data', [])
        
        if not model_type or not data:
            return Response({'error': 'Model type and data are required'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        try:
            app_label, model_name = model_type.split('.')
            model = apps.get_model(app_label, model_name)
        except Exception:
            return Response({'error': f'Invalid model type: {model_type}'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        # Log the operation
        operation_log = BulkOperationLog.objects.create(
            user=request.user,
            operation_type='bulk_create',
            model_type=model_type,
            records_count=len(data),
            status='in_progress'
        )
        
        try:
            created_objects = []
            errors = []
            
            with transaction.atomic():
                for item_data in data:
                    try:
                        obj = model.objects.create(**item_data)
                        created_objects.append(obj)
                    except Exception as e:
                        errors.append({
                            'data': item_data,
                            'error': str(e)
                        })
            
            # Update operation log
            operation_log.status = 'completed'
            operation_log.success_count = len(created_objects)
            operation_log.error_count = len(errors)
            operation_log.save()
            
            return Response({
                'message': f'Successfully created {len(created_objects)} records',
                'created_count': len(created_objects),
                'error_count': len(errors),
                'errors': errors,
                'operation_id': operation_log.id
            })
            
        except Exception as e:
            operation_log.status = 'failed'
            operation_log.error_message = str(e)
            operation_log.save()
            
            return Response({'error': f'Bulk operation failed: {str(e)}'}, 
                          status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['put'])
    def bulk_update(self, request):
        """Bulk update records"""
        model_type = request.data.get('model_type')
        updates = request.data.get('updates', [])
        
        if not model_type or not updates:
            return Response({'error': 'Model type and updates are required'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        try:
            app_label, model_name = model_type.split('.')
            model = apps.get_model(app_label, model_name)
        except Exception:
            return Response({'error': f'Invalid model type: {model_type}'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        # Log the operation
        operation_log = BulkOperationLog.objects.create(
            user=request.user,
            operation_type='bulk_update',
            model_type=model_type,
            records_count=len(updates),
            status='in_progress'
        )
        
        try:
            updated_count = 0
            errors = []
            
            with transaction.atomic():
                for update_data in updates:
                    try:
                        record_id = update_data.pop('id')
                        obj = model.objects.get(id=record_id)
                        
                        for field, value in update_data.items():
                            setattr(obj, field, value)
                        
                        obj.save()
                        updated_count += 1
                        
                    except Exception as e:
                        errors.append({
                            'id': update_data.get('id'),
                            'error': str(e)
                        })
            
            # Update operation log
            operation_log.status = 'completed'
            operation_log.success_count = updated_count
            operation_log.error_count = len(errors)
            operation_log.save()
            
            return Response({
                'message': f'Successfully updated {updated_count} records',
                'updated_count': updated_count,
                'error_count': len(errors),
                'errors': errors,
                'operation_id': operation_log.id
            })
            
        except Exception as e:
            operation_log.status = 'failed'
            operation_log.error_message = str(e)
            operation_log.save()
            
            return Response({'error': f'Bulk operation failed: {str(e)}'}, 
                          status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['delete'])
    def bulk_delete(self, request):
        """Bulk delete records"""
        model_type = request.data.get('model_type')
        record_ids = request.data.get('record_ids', [])
        
        if not model_type or not record_ids:
            return Response({'error': 'Model type and record IDs are required'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        try:
            app_label, model_name = model_type.split('.')
            model = apps.get_model(app_label, model_name)
        except Exception:
            return Response({'error': f'Invalid model type: {model_type}'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        # Log the operation
        operation_log = BulkOperationLog.objects.create(
            user=request.user,
            operation_type='bulk_delete',
            model_type=model_type,
            records_count=len(record_ids),
            status='in_progress'
        )
        
        try:
            deleted_count = 0
            errors = []
            
            with transaction.atomic():
                for record_id in record_ids:
                    try:
                        obj = model.objects.get(id=record_id)
                        obj.delete()
                        deleted_count += 1
                        
                    except Exception as e:
                        errors.append({
                            'id': record_id,
                            'error': str(e)
                        })
            
            # Update operation log
            operation_log.status = 'completed'
            operation_log.success_count = deleted_count
            operation_log.error_count = len(errors)
            operation_log.save()
            
            return Response({
                'message': f'Successfully deleted {deleted_count} records',
                'deleted_count': deleted_count,
                'error_count': len(errors),
                'errors': errors,
                'operation_id': operation_log.id
            })
            
        except Exception as e:
            operation_log.status = 'failed'
            operation_log.error_message = str(e)
            operation_log.save()
            
            return Response({'error': f'Bulk operation failed: {str(e)}'}, 
                          status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ============================================================================
# EXPORT/IMPORT VIEWSET
# ============================================================================

class ExportImportViewSet(viewsets.ViewSet):
    """ViewSet for data export and import operations"""
    
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FileUploadParser]
    
    @action(detail=False, methods=['get'])
    def export(self, request):
        """Export data in various formats"""
        model_type = request.query_params.get('model_type')
        format_type = request.query_params.get('format', 'csv')
        filters = request.query_params.get('filters', '{}')
        
        if not model_type:
            return Response({'error': 'Model type is required'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        try:
            app_label, model_name = model_type.split('.')
            model = apps.get_model(app_label, model_name)
        except Exception:
            return Response({'error': f'Invalid model type: {model_type}'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Parse filters
            filter_dict = json.loads(filters)
            queryset = model.objects.filter(**filter_dict)
            
            if format_type == 'csv':
                response = HttpResponse(content_type='text/csv')
                response['Content-Disposition'] = f'attachment; filename="{model_name}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
                
                writer = csv.writer(response)
                
                # Write headers
                field_names = [field.name for field in model._meta.fields]
                writer.writerow(field_names)
                
                # Write data
                for obj in queryset:
                    row = []
                    for field_name in field_names:
                        value = getattr(obj, field_name)
                        row.append(str(value) if value is not None else '')
                    writer.writerow(row)
                
                return response
                
            elif format_type == 'json':
                data = list(queryset.values())
                return JsonResponse(data, safe=False)
                
            elif format_type == 'excel':
                response = HttpResponse(
                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )
                response['Content-Disposition'] = f'attachment; filename="{model_name}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
                
                wb = Workbook()
                ws = wb.active
                ws.title = model_name
                
                # Write headers
                field_names = [field.name for field in model._meta.fields]
                for col, field_name in enumerate(field_names, 1):
                    ws.cell(row=1, column=col, value=field_name)
                
                # Write data
                for row, obj in enumerate(queryset, 2):
                    for col, field_name in enumerate(field_names, 1):
                        value = getattr(obj, field_name)
                        ws.cell(row=row, column=col, value=str(value) if value is not None else '')
                
                wb.save(response)
                return response
                
            else:
                return Response({'error': f'Unsupported format: {format_type}'}, 
                              status=status.HTTP_400_BAD_REQUEST)
                
        except Exception as e:
            return Response({'error': f'Export failed: {str(e)}'}, 
                          status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['post'])
    def import_data(self, request):
        """Import data from file"""
        model_type = request.data.get('model_type')
        file = request.FILES.get('file')
        format_type = request.data.get('format', 'csv')
        
        if not model_type or not file:
            return Response({'error': 'Model type and file are required'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        try:
            app_label, model_name = model_type.split('.')
            model = apps.get_model(app_label, model_name)
        except Exception:
            return Response({'error': f'Invalid model type: {model_type}'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        try:
            imported_count = 0
            errors = []
            
            if format_type == 'csv':
                # Read CSV file
                decoded_file = file.read().decode('utf-8')
                csv_data = csv.DictReader(io.StringIO(decoded_file))
                
                with transaction.atomic():
                    for row in csv_data:
                        try:
                            # Clean and validate data
                            cleaned_data = {}
                            for field_name, value in row.items():
                                if field_name in [f.name for f in model._meta.fields]:
                                    cleaned_data[field_name] = value if value != '' else None
                            
                            obj = model.objects.create(**cleaned_data)
                            imported_count += 1
                            
                        except Exception as e:
                            errors.append({
                                'row': row,
                                'error': str(e)
                            })
                            
            elif format_type == 'json':
                # Read JSON file
                data = json.loads(file.read().decode('utf-8'))
                
                with transaction.atomic():
                    for item in data:
                        try:
                            obj = model.objects.create(**item)
                            imported_count += 1
                            
                        except Exception as e:
                            errors.append({
                                'item': item,
                                'error': str(e)
                            })
                            
            elif format_type == 'excel':
                # Read Excel file
                wb = load_workbook(file)
                ws = wb.active
                
                # Get headers from first row
                headers = []
                for col in range(1, ws.max_column + 1):
                    headers.append(ws.cell(row=1, column=col).value)
                
                with transaction.atomic():
                    for row in range(2, ws.max_row + 1):
                        try:
                            row_data = {}
                            for col, header in enumerate(headers, 1):
                                if header in [f.name for f in model._meta.fields]:
                                    value = ws.cell(row=row, column=col).value
                                    row_data[header] = str(value) if value is not None else None
                            
                            obj = model.objects.create(**row_data)
                            imported_count += 1
                            
                        except Exception as e:
                            errors.append({
                                'row': row,
                                'error': str(e)
                            })
                            
            else:
                return Response({'error': f'Unsupported format: {format_type}'}, 
                              status=status.HTTP_400_BAD_REQUEST)
            
            return Response({
                'message': f'Successfully imported {imported_count} records',
                'imported_count': imported_count,
                'error_count': len(errors),
                'errors': errors
            })
            
        except Exception as e:
            return Response({'error': f'Import failed: {str(e)}'}, 
                          status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ============================================================================
# GLOBAL STATISTICS VIEWSET
# ============================================================================

class GlobalStatisticsViewSet(viewsets.ViewSet):
    """ViewSet for global system statistics and analytics"""
    
    permission_classes = [IsAuthenticated]
    
    @action(detail=False, methods=['get'])
    def overview(self, request):
        """Get overview statistics for the entire system"""
        try:
            # Get model counts
            models_to_count = [
                'api.User', 'api.Customer', 'api.Pharmacy', 'api.Rider',
                'api.MedicineCatalog', 'api.Order', 'api.Payment'
            ]
            
            model_counts = {}
            for model_path in models_to_count:
                try:
                    app_label, model_name = model_path.split('.')
                    model = apps.get_model(app_label, model_name)
                    model_counts[model_name.lower()] = model.objects.count()
                except Exception:
                    model_counts[model_name.lower()] = 0
            
            # Get recent activity
            recent_orders = apps.get_model('api', 'Order').objects.filter(
                created_at__gte=timezone.now() - timedelta(days=7)
            ).count()
            
            recent_users = apps.get_model('api', 'User').objects.filter(
                date_joined__gte=timezone.now() - timedelta(days=7)
            ).count()
            
            # Get system health status
            try:
                latest_health = SystemHealth.objects.filter(
                    component='overall'
                ).latest('last_check')
                system_status = latest_health.status
            except SystemHealth.DoesNotExist:
                system_status = 'unknown'
            
            # Get API usage stats
            api_requests_today = ApiUsage.objects.filter(
                timestamp__date=timezone.now().date()
            ).count()
            
            api_requests_week = ApiUsage.objects.filter(
                timestamp__gte=timezone.now() - timedelta(days=7)
            ).count()
            
            overview_data = {
                'model_counts': model_counts,
                'recent_activity': {
                    'orders_last_7_days': recent_orders,
                    'new_users_last_7_days': recent_users
                },
                'system_status': system_status,
                'api_usage': {
                    'requests_today': api_requests_today,
                    'requests_last_7_days': api_requests_week
                },
                'timestamp': timezone.now().isoformat()
            }
            
            return Response(overview_data)
            
        except Exception as e:
            return Response({'error': f'Failed to get overview: {str(e)}'}, 
                          status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'])
    def model_stats(self, request):
        """Get detailed statistics for a specific model"""
        model_type = request.query_params.get('model_type')
        
        if not model_type:
            return Response({'error': 'Model type is required'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        try:
            app_label, model_name = model_type.split('.')
            model = apps.get_model(app_label, model_name)
        except Exception:
            return Response({'error': f'Invalid model type: {model_type}'}, 
                          status=status.HTTP_400_BAD_REQUEST)
        
        try:
            # Get basic counts
            total_count = model.objects.count()
            
            # Get creation trends (last 30 days)
            creation_trends = []
            for i in range(30):
                date = timezone.now().date() - timedelta(days=i)
                count = model.objects.filter(
                    created_at__date=date
                ).count()
                creation_trends.append({
                    'date': date.isoformat(),
                    'count': count
                })
            
            # Get field statistics if applicable
            field_stats = {}
            for field in model._meta.fields:
                if field.get_internal_type() in ['CharField', 'TextField']:
                    # Get unique values count
                    unique_count = model.objects.values(field.name).distinct().count()
                    field_stats[field.name] = {
                        'unique_values': unique_count,
                        'field_type': field.get_internal_type()
                    }
            
            model_stats_data = {
                'model_name': model_name,
                'total_count': total_count,
                'creation_trends': creation_trends,
                'field_statistics': field_stats,
                'timestamp': timezone.now().isoformat()
            }
            
            return Response(model_stats_data)
            
        except Exception as e:
            return Response({'error': f'Failed to get model stats: {str(e)}'}, 
                          status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'])
    def performance(self, request):
        """Get system performance metrics"""
        try:
            # Database performance
            with connection.cursor() as cursor:
                cursor.execute("SELECT version()")
                db_version = cursor.fetchone()[0]
            
            # Cache performance
            try:
                from django.core.cache import cache
                start_time = time.time()
                cache.set('perf_test', 'test_value', 10)
                cache.get('perf_test')
                cache_response_time = (time.time() - start_time) * 1000  # Convert to milliseconds
                cache_status = 'healthy'
            except Exception:
                cache_response_time = None
                cache_status = 'unavailable'
            
            # API response times
            recent_api_usage = ApiUsage.objects.filter(
                timestamp__gte=timezone.now() - timedelta(hours=1)
            ).aggregate(
                avg_response_time=Avg('response_time'),
                max_response_time=Max('response_time'),
                min_response_time=Min('response_time')
            )
            
            performance_data = {
                'database': {
                    'version': db_version,
                    'status': 'healthy'
                },
                'cache': {
                    'status': cache_status,
                    'response_time_ms': cache_response_time
                },
                'api_performance': {
                    'average_response_time_ms': recent_api_usage['avg_response_time'] or 0,
                    'max_response_time_ms': recent_api_usage['max_response_time'] or 0,
                    'min_response_time_ms': recent_api_usage['min_response_time'] or 0
                },
                'timestamp': timezone.now().isoformat()
            }
            
            return Response(performance_data)
            
        except Exception as e:
            return Response({'error': f'Failed to get performance metrics: {str(e)}'}, 
                          status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# ============================================================================
# BULK OPERATION LOG VIEWSET
# ============================================================================

class BulkOperationLogViewSet(viewsets.ReadOnlyModelViewSet):
    """ViewSet for viewing bulk operation logs"""
    
    queryset = BulkOperationLog.objects.all()
    permission_classes = [IsAuthenticated]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['operation_type', 'model_type', 'status', 'user']
    search_fields = ['operation_type', 'model_type', 'error_message']
    ordering_fields = ['timestamp', 'records_count', 'success_count', 'error_count']
    ordering = ['-timestamp']
    
    def get_queryset(self):
        """Filter queryset based on user permissions"""
        if self.request.user.is_staff:
            return BulkOperationLog.objects.all()
        return BulkOperationLog.objects.filter(user=self.request.user)
